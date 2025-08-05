# -*- coding:utf-8 -*-
import time
import datetime
import pickle
import numpy as np
import openpyxl
from sklearn import preprocessing
from model import *
from train_chunk1 import *
from tools import *

os.environ["CUDA_VISIBLE_DEVICES"] = "0"


def computer_fisher(model, imgset, num_sample=30):
    """Compute the Fisher information for model parameters using a sample of input data."""
    f_accum = []
    
    # Initialize f_accum[i] as a zero tensor with the same shape as the corresponding weight tensor
    for i in range(len(model.weights)):
        f_accum.append(tf.zeros_like(model.weights[i]))

    with tf.GradientTape(persistent=True) as tape:
        tape.watch(model.weights)
        for j in range(num_sample):
            img_index = np.random.randint(imgset.shape[0])
            output = model(np.expand_dims(imgset[img_index], 0))
            log_output = tf.math.log(output)
            grads = tape.gradient(log_output, model.weights)
            
            # Square each gradient tensor and accumulate it in f_accum[i]
            for i, grad in enumerate(grads):
                f_accum[i] += tf.square(grad)

    # Normalize each tensor in f_accum by dividing by the number of samples
    f_accum = [f / num_sample for f in f_accum]
    return f_accum


#################################################
######            Initialization           ######
#################################################
data_folder = "./data"
data_folder_test = "./data/wave/test/"
trained_model_folder = "./train/"
model_folder = "./model/"
final_folder = "./result/"

nb_epochs = 30
batch_size = 16
input_size = 450
learning_rate = 5e-4

num = 256           # Total number of chunks to process
num_s = 1           # Number of keys to recover simultaneously in each chunk
n_tra = 35          # Number of training traces
n_val = 15          # Number of validation traces
point = input_size  # Number of points in each trace
j = n_tra + n_val   # Used for normalization
group_sum = 10      # Run for 10 groups and take the average
lambda_mmd = 5      # Weight for the MMD loss

start_time = time.time()

book = openpyxl.Workbook()

sheet = book.active
sheet.title = 'Sheet'

# Write the table header to the file
header = ['Group'] + [f's[{i}]' for i in range(256)]

# The key is only used for final accuracy verification.
keys = load_key("Kyber768_key.txt", start=0, length=256)

# data contains traces, and each line in wave.txt stores the trace corresponding to 256 keys.
data = np.loadtxt(data_folder + "/wave/wave_sk0.txt")

# label is the message bits corresponding to the coefficients of the secret key with new chosen ciphertexts.
labels = np.loadtxt(data_folder + '/label/guessM_sk0.txt')

for i in range(len(header)):
    sheet.cell(row=1, column=i+1, value=header[i])

for i in range(len(keys)):
    sheet.cell(row=2, column=i+1, value=keys[i])

for i in range(group_sum):
    sheet.cell(row=i+3, column=1, value="Group " + str(i + 1))

for group in range(group_sum):
    # source domain
    temp = data[j * group:j * (group + 1), 0:point]
    np.savetxt(data_folder_test + './new_data.txt', temp)
    data_new = np.loadtxt(data_folder_test + './new_data.txt')
    np.save(data_folder_test + './new_data.npy', data_new)

    data_array = []
    for i in range(5):
        label = []
        for row in range(j * group, (group + 1) * j):
            label.append(labels[row, i])
        np.save(data_folder_test + './label_' + str(i - 2) + '.npy', label)
        np.save(data_folder_test + './label_sum.npy', label)

        Y_profiling = np.load(data_folder_test + './label_sum.npy')
        X_profiling = np.load(data_folder_test + './new_data.npy')
        X_profiling1 = X_profiling
        Y_profiling1 = Y_profiling
        (X_profiling, Y_profiling) = shuffle_data(X_profiling, Y_profiling)
        X_profiling = X_profiling.astype('float32')
        X_profiling1 = X_profiling1.astype('float32')

        # Apply standard normalization
        scaler = preprocessing.StandardScaler()
        X_profiling = scaler.fit_transform(X_profiling)
        X_profiling1 = scaler.fit_transform(X_profiling1)
        
        # Apply MinMax scaling to [0, 1]
        scaler = preprocessing.MinMaxScaler(feature_range=(0, 1))
        X_profiling = scaler.fit_transform(X_profiling)
        X_profiling1 = scaler.fit_transform(X_profiling1)

        # Record the metrics
        model = cnn_architecture(input_size=input_size, classes=2)
        history = train_model(i - 2, X_profiling[:n_tra], Y_profiling[:n_tra],
                              X_profiling[n_tra:n_val + n_tra], Y_profiling[n_tra:n_val + n_tra],
                              model, trained_model_folder, learning_rate=learning_rate,
                              epochs=nb_epochs, batch_size=batch_size)

        model_name = "model_weights_" + str(i - 2)
        with open(model_folder + 'history_' + model_name, 'wb') as file_pi:
            pickle.dump(history.history, file_pi)

        if not os.path.exists(final_folder):
            os.makedirs(final_folder)

        newfolder = final_folder + 'group-' + str(group + 1) + '/' + 'sk 1/'
        if not os.path.exists(newfolder):
            os.makedirs(newfolder)
        floder = newfolder

        save_chunk1_history(history, 'loss', floder, 100, 0, 10, i - 2)
        save_chunk1_history(history, 'accuracy', floder, 100, 0, 10, i - 2)

        # Evaluate using Pearson correlation
        label1 = Y_profiling1
        predictions = model.predict(X_profiling1)
        label2 = np.argmax(predictions, axis=1)

        # Compute Pearson correlation and final loss
        correlation = pearson_correlation(label1, label2)
        final_loss = 1 - correlation
        data_array.append(final_loss)

    # Find the correct key
    data_tensor = tf.convert_to_tensor(data_array)
    min_value = tf.reduce_min(data_tensor)
    min_index = tf.argmin(data_tensor).numpy() - 2
    sheet.cell(row=group + 3, column=2).value = int(min_index)

    # Load the pretrained model using the best s index
    label_y_corr = np.load(data_folder_test + './label_' + str(min_index) + '.npy')
    model = tf.keras.models.load_model(trained_model_folder + 'model_weights_' + str(min_index) + '.h5')
    
    print('Processing Fisher Information...')
    # Compute Fisher Information matrix for EWC regularization
    I = computer_fisher(model, X_profiling1[:2 * n_tra])
    print('Processing Finish!')
    
    # Build a new model with EWC support and load pretrained weights
    model = cnn_architecture_EWC(I=I, input_size=input_size, classes=2, pretrained_model=model)
    
    # Load weights again into the new EWC model architecture
    model.load_weights(trained_model_folder + 'model_weights_' + str(min_index) + '.h5')

    for m in range(num - 1):
        # source domain
        temp = data[j * group:j * (group + 1), m * point:(m + 1) * point]
        np.savetxt(data_folder_test + './new_data.txt', temp)
        data_new = np.loadtxt(data_folder_test + './new_data.txt')
        np.save(data_folder_test + './new_data.npy', data_new)

        # target domain
        temp_1 = data[j * group:j * (group + 1), (m + 1) * point:(m + 2) * point]
        np.savetxt(data_folder_test + './new_data_1.txt', temp_1)
        data_1 = np.loadtxt(data_folder_test + './new_data_1.txt')
        np.save(data_folder_test + './new_data_1.npy', data_1)

        data_y = np.load(data_folder_test + './new_data.npy')
        data_t = np.load(data_folder_test + './new_data_1.npy')
        label_y = label_y_corr
        label_t = label_y_corr
        data_t1 = data_t
        data_y, label_y, data_t, q1 = shuffle_data_4(data_y, label_y, data_t, label_t)

        # Standardization and Normalization (between 0 and 1)
        scaler = preprocessing.StandardScaler()
        data_y = scaler.fit_transform(data_y)
        data_t = scaler.fit_transform(data_t)
        data_t1 = scaler.fit_transform(data_t1)
        scaler = preprocessing.MinMaxScaler(feature_range=(0, 1))
        data_y = scaler.fit_transform(data_y)
        data_t = scaler.fit_transform(data_t)
        data_t1 = scaler.fit_transform(data_t1)

        newfolder = final_folder + 'group-' + str(group + 1) + '/' + 'sk ' + str(m + 2) + '/'
        if not os.path.exists(newfolder):
            os.makedirs(newfolder)
        floder = newfolder

        # Train model with source and target domain data using MMD-based domain adaptation
        history = train_model_st(data_y[:n_tra], label_y[:n_tra], data_t[:n_tra],
                                 data_y[n_tra:n_val + n_tra], label_y[n_tra:n_val + n_tra], model,
                                 trained_model_folder, floder, epochs=nb_epochs, batch_size=batch_size,
                                 learning_rate=learning_rate, lambda_mmd=lambda_mmd, num_classes=2)

        # Predict labels for target domain
        predictions = model.predict(data_t1)
        predicted_labels = np.argmax(predictions, axis=1)

        val_loss_dic = {}
        min_index = -1
        min_value = float('inf')
        for i in range(5):
            label_guess = []
            for row in range(j * group, (group + 1) * j):
                label_guess.append(labels[row, i + (m + 1) * 5])
            
            np.save(data_folder_test + './label_sum_' + str(i - 2) + '.npy', label_guess)
            np.save(data_folder_test + './label_sum.npy', label_guess)

            label_guess = np.load(data_folder_test + './label_sum.npy')

            correlation = pearson_correlation(label_guess, predicted_labels)
            final_loss = 1 - correlation
            val_loss_dic[len(val_loss_dic) + 1] = final_loss

        # Find the best key with the lowest validation loss
        min_value = min(val_loss_dic.values())
        min_index = min(val_loss_dic, key=val_loss_dic.get)
        sn_min_index = min_index - 3

        print("Group " + str(group) + ": Key " + str(m + 1))
        print("Validation Loss Dictionary: ", val_loss_dic)
        print("Minimum Value: ", min_value)
        print("Guessed Key for sn: ", sn_min_index)

        sheet.cell(row=group + 3, column=m + 3).value = sn_min_index

        I = computer_fisher(model, X_profiling1[:2 * n_tra])
        model = cnn_architecture_EWC(I=I, input_size=input_size, classes=2, pretrained_model=model)
        label_y_corr = np.load(data_folder_test + './label_sum_' + str(sn_min_index) + '.npy')
        model.load_weights(trained_model_folder + 'model_weights_mmd1.h5')

    book.save(final_folder + '/sk_sum.xlsx')

    
# Compute the prediction probability of each individual key
wb = openpyxl.load_workbook(final_folder + '/sk_sum.xlsx')
sheet = wb['Sheet']

sheet.cell(row=group_sum + 3, column=1, value="Accuracy")

for j in range(2, num + 2):
    sum1 = 0
    count = 0
    flag = sheet.cell(row=2, column=j).value
    for i in range(3, 3 + group_sum):
        if sheet.cell(row=i, column=j).value == flag:
            sum1 = sum1 + 1
        count = count + 1
    value = sum1 / count
    sheet.cell(row=group_sum + 3, column=j, value=value)

wb.save(final_folder + 'sk_sum.xlsx')

# Compute the overall prediction probability across all keys
wb = openpyxl.load_workbook(final_folder + 'sk_sum.xlsx')
sheet = wb['Sheet']

accuracy_row = group_sum + 3

accuracies = []
for col in range(2, 2 + num):
    cell_value = sheet.cell(row=accuracy_row, column=col).value
    if isinstance(cell_value, (int, float)):
        accuracies.append(cell_value)

average_accuracy = sum(accuracies) / len(accuracies)

print("#############################################")
print("Average accuracy: ", average_accuracy)
sheet.cell(row=accuracy_row + 1, column=1, value="Average Accuracy")
sheet.cell(row=accuracy_row + 1, column=2, value=round(average_accuracy, 4))

wb.save(final_folder + 'sk_sum.xlsx')

end_time = time.time()
duration = end_time - start_time

hours, rem = divmod(duration, 3600)
minutes, seconds = divmod(rem, 60)

print("block=2, total groups: " + str(group_sum))
print("Start time:", datetime.datetime.fromtimestamp(start_time).strftime('%Y-%m-%d %H:%M:%S'))
print("End time:", datetime.datetime.fromtimestamp(end_time).strftime('%Y-%m-%d %H:%M:%S'))
print("Total duration: {:0>2}:{:0>2}:{:05.2f}".format(int(hours), int(minutes), seconds))

